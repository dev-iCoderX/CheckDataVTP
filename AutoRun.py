import tkinter as tk
from tkinter import *
from openpyxl import load_workbook as Workbook
from selenium import webdriver
import tkinter.messagebox as Msb
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from threading import Thread
from time import sleep


class AutoRun(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.parent = parent
        self.parent.title("Auto Crap")
        img = tk.Image("photo", file="./images/icon.png")
        self.parent.iconbitmap(r'./images/icon.ico')
        self.parent.resizable(0, 0)
        self.parent.tk.call('wm', 'iconphoto', self.parent._w, img)
        self.lockStatus = True
        self.lowKey = "abcdefghijklmnopqrstuvwxyz"
        self.Confirm()

    def Confirm(self):
        self.frame1 = Frame(
            self.parent, relief=RAISED, borderwidth=1)
        self.frame1.pack(fill=X, padx=0, pady=0)

        self.value_input_start = StringVar(self.parent, value="")
        self.input_start = Entry(
            self.frame1, width=20, textvariable=self.value_input_start, show='*')
        self.input_start.pack(side=LEFT, padx=5, pady=5)

        self.login_start = Button(
            self.frame1, text="Đăng nhập", width=10, command=self.ConfirmFuntion)
        self.login_start.pack(side=LEFT, padx=5, pady=5)

    def ConfirmFuntion(self):
        # if(self.input_start.get() == "duykhanh111"):
        if(self.input_start.get() == ""):
            self.Forget()
            self.initUI()
            self.lockStatus = False
        else:
            Msb.showwarning(title='Cảnh báo', message='Hãy nhập đúng mật khẩu')

    def Forget(self):
        self.frame1.forget()
        try:
            self.frame2.forget()
            self.frame3.forget()
            self.frame5.forget()
        except:
            pass
        #self.input_start.forget()
        #self.login_start.forget()

    def initUI(self):
        self.frame1 = Frame(
            self.parent, relief=RAISED, borderwidth=1)
        self.frame1.pack(fill=X, padx=0, pady=0)

        self.start_line_label = Label(
            self.frame1, text="Bắt đầu từ dòng: ", width= 15, anchor='w')
        self.start_line_label.pack(side=LEFT, padx=5, pady=5)

        self.value_input_start = StringVar(self.parent, value='2')
        self.input_start = Entry(
            self.frame1, width=15, textvariable=self.value_input_start)
        self.input_start.pack(side=LEFT, padx=5, pady=5)

        self.start_line_label = Label(
            self.frame1, text="Bắt đầu từ cột: ", width= 15, anchor='w')
        self.start_line_label.pack(side=LEFT, padx=5, pady=5)

        self.value_input_col_start = StringVar(self.parent, value='a')
        self.input_col_start = Entry(
            self.frame1, width=15, textvariable=self.value_input_col_start)
        self.input_col_start.pack(side=LEFT, padx=5, pady=5)

        self.frame5 = Frame(
            self.parent, relief=RAISED, borderwidth=1)
        self.frame5.pack(fill=X, padx=0, pady=0)

        self.end_line_label = Label(
            self.frame5, text="Kết thúc dòng: ", width= 15, anchor='w')
        self.end_line_label.pack(side=LEFT, padx=5, pady=5)

        self.value_input_end = StringVar(self.parent, value='10')
        self.input_end = Entry(
            self.frame5, width=15, textvariable=self.value_input_end)
        self.input_end.pack(side=LEFT, padx=5, pady=5)

        self.end_line_label = Label(
            self.frame5, text="Kết thúc cột: ", width= 15, anchor='w')
        self.end_line_label.pack(side=LEFT, padx=5, pady=5)

        self.value_input_col_end = StringVar(self.parent, value='e')
        self.input_col_end = Entry(
            self.frame5, width=15, textvariable=self.value_input_col_end)
        self.input_col_end.pack(side=LEFT, padx=5, pady=5)

        self.frame3 = Frame(
            self.parent, relief=RAISED, borderwidth=1)
        self.frame3.pack(fill=X, padx=0, pady=0)

        login_button = Button(self.frame3, text="Đăng nhập",
                              width=10, command=self.Login)
        login_button.pack(side=RIGHT, padx=5, pady=5)

        login_start = Button(self.frame3, text="Bắt đầu",
                             width=10, command=self.Start)
        login_start.pack(side=RIGHT, padx=5, pady=5)

        lock_button = Button(self.frame3, text="Khóa",
                              width=10, command=self.Lock)
        lock_button.pack(side=RIGHT, padx=5, pady=5)

        self.frame2 = Frame(
            self.parent, relief=RAISED, borderwidth=1)
        self.frame2.pack(fill=X, padx=0, pady=0)

        self.count_label = Label(
            self.frame2, text="Đã làm: 0", anchor='w', fg="green")
        self.count_label.pack(side=LEFT, padx=5, pady=0)

        self.count_left_label = Label(
            self.frame2, text="Còn: 0", anchor='w', fg="red")
        self.count_left_label.pack(side=LEFT, padx=5, pady=0)

        self.status_label = Label(
            self.frame2, text="Trạng thái: Nghỉ", anchor='w', fg="purple")
        self.status_label.pack(side=RIGHT, padx=5, pady=0)

    def Lock(self):
        self.Forget()
        self.Confirm()
        self.lockStatus = True

    def Login(self):
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.driver = webdriver.Chrome("chromedriver.exe", options=options)
        self.driver.get("https://evtp2.viettelpost.vn/?uId=tra-cuu-hanh-trinh")

    def Run(self):
        self.status_label["text"] = "Trạng thái: Đang chạy"
        self.status_label.config(fg="green")
        self.count_label["text"] = "Đã làm: 0"
        self.count_left_label["text"] = "Còn: 0"
        try:
            self.driver.get(
                "https://evtp2.viettelpost.vn/?uId=tra-cuu-hanh-trinh")
            book = Workbook('data.xlsx')
            sheet = book.active
            row = sheet.max_row
            charFirst = self.input_col_start.get()
            charLast = self.input_col_end.get()
            location = self.lowKey.find(charFirst)
            locationLast = self.lowKey.find(charLast)
            firstRow = self.input_start.get()
            lastRow = self.input_end.get()
            count = 0
            countLeft = int(self.input_end.get()) - int(self.input_start.get()) + 1
            self.count_left_label["text"] = "Còn: " + str(countLeft)
            preData = [1,1,1,1,1]
            oldData = []
            preKey = ""
            for row in range(int(firstRow),int(lastRow)+1):
                rawData = []
                count += 1
                countLeft -= 1
                countIn = 1
                for cell in range(len(sheet[row])):
                    if(cell == location):
                        while True:
                            print(sheet[row][cell].value)
                            inputLabel = self.driver.find_element_by_class_name(
                                "z-bandbox-input")
                            ActionChains(self.driver).click(inputLabel).key_down(Keys.CONTROL).send_keys(
                                "a").key_up(Keys.CONTROL).send_keys(sheet[row][cell].value).perform()
                            inputLabel.send_keys(Keys.ENTER)
                            rawData.append(sheet[row][cell].value)
                            oldData = self.GetData()
                            rawData += oldData
                            print (rawData)
                            if(oldData == preData and preKey != sheet[row][cell].value):
                                rawData = []
                                print("Success")
                            else:
                                preData = oldData
                                preKey = sheet[row][cell].value
                                break
                    elif(cell > location):
                        sheet[row][cell].value = rawData[countIn]
                        countIn += 1
                if(not self.lockStatus):
                    self.count_label["text"] = "Đã làm: " + str(count)
                    self.count_left_label["text"] = "Còn: " + str(countLeft)
            book.save('new_data.xlsx')
            self.status_label["text"] = "Trạng thái: Hoàn thành"
            self.status_label.config(fg="blue")
            Msb.showinfo(title='Thông báo', message='Đã hoàn thành công việc')
            self.status_label["text"] = "Trạng thái: Nghỉ"
            self.status_label.config(fg="purple")
        except Exception as e:
            book.save('new_data.xlsx')
            self.status_label["text"] = "Trạng thái: Lỗi"
            self.status_label.config(fg="red")
            Msb.showerror(title='Cảnh báo', message=e)
            self.status_label["text"] = "Trạng thái: Nghỉ"
            self.status_label.config(fg="purple")
            self.driver.close()

    def GetData(self):
        oldData = []
        count = 200000
        while count > 0:
            try:
                oldData.append(self.driver.find_element_by_xpath(
                    "/html/body/div[1]/div[1]/div/div[2]/div/div[2]/div/div[2]/div/div/div/div/div[2]/div[1]/div/div[2]/div[2]/span[2]").text)
                oldData.append(self.driver.find_element_by_xpath(
                    "/html/body/div[1]/div[1]/div/div[2]/div/div[2]/div/div[2]/div/div/div/div/div[2]/div[1]/div/div[4]/div[2]/span[2]").text)
                oldData.append(self.driver.find_element_by_xpath(
                    "/html/body/div[1]/div[1]/div/div[2]/div/div[2]/div/div[2]/div/div/div/div/div[2]/div[1]/div/div[3]/div[2]/span[2]").text)
                oldData.append(self.driver.find_element_by_xpath(
                    "/html/body/div[1]/div[1]/div/div[2]/div/div[2]/div/div[2]/div/div/div/div/div[2]/div[1]/div/div[2]/div[1]/span[2]").text)
                break
            except:
                oldData = []
                print("Fail")
            count -= 1
        return oldData

    def Start(self):
        self.t = Thread(target=self.Run).start()
