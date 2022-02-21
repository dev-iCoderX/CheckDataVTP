from openpyxl import load_workbook as Workbook
book = Workbook('data.xlsx')
sheet = book.active
row = sheet.max_row
firstRow = 'b2'
lastRow = 'f' + str(row)
for row in sheet[firstRow:lastRow]:
    rawData = []
    for index, cell in enumerate(row):
        if(index == 0):
            rawData.append(cell.value)
            rawData == [1,1,1,1,1]
        else:
            cell.value = rawData[index]
book.save('new_data.xlsx')